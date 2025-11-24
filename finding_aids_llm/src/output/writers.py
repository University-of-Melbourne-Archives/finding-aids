# src/output/writers.py
from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Dict, Iterable, List

try:
    import pandas as pd  # type: ignore[import]
except ImportError:  # pandas is optional
    pd = None

import csv




# Scalar fields to flatten into *_value / *_confidence columns
SCALAR_FIELDS = [
    "group",
    "group_notes",
    "series",
    "series_notes",
    "unit",
    "finding_aid_reference_raw",
    "text",
    "start_date_original",
    "end_date_original",
    "start_date_formatted",
    "end_date_formatted",
]


def write_items_json(path: Path, items: List[Dict[str, Any]]) -> None:
    """
    Write the final combined items to a JSON file with shape:

        { "items": [ ... ] }
    """
    obj = {"items": items}
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(obj, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def write_issues_json(path: Path, issues: Iterable[Dict[str, Any]]) -> None:
    """
    Write parse issues (each a dict) to JSON for logging/debugging.
    """
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(list(issues), ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def _flatten_item_for_tabular(item: Dict[str, Any]) -> Dict[str, Any]:
    """
    Flatten one item into a single row suitable for CSV/XLSX.

    - page.chunk -> page_chunk
    - page.page_number -> page_number
    - scalar fields (e.g. group) -> group_value, group_confidence
    - annotations -> JSON string
    """
    row: Dict[str, Any] = {}

    page = item.get("page") or {}
    if isinstance(page, dict):
        row["page_chunk"] = page.get("chunk")
        row["page_number"] = page.get("page_number")
    else:
        row["page_chunk"] = None
        row["page_number"] = None

    for field in SCALAR_FIELDS:
        val = item.get(field)
        if isinstance(val, dict):
            row[f"{field}_value"] = val.get("value")
            row[f"{field}_confidence"] = val.get("confidence")
        else:
            row[f"{field}_value"] = None
            row[f"{field}_confidence"] = None

    # annotations as JSON string
    anns = item.get("annotations", [])
    row["annotations"] = json.dumps(anns, ensure_ascii=False)

    return row


def write_items_csv(path: Path, items: List[Dict[str, Any]]) -> None:
    """
    Write combined items to CSV. Uses pandas if available,
    otherwise falls back to csv.DictWriter.
    """
    rows = [_flatten_item_for_tabular(it) for it in items]
    path.parent.mkdir(parents=True, exist_ok=True)

    if not rows:
        # create empty file with no rows; header depends on pandas/CSV
        if pd is not None:
            pd.DataFrame([]).to_csv(path, index=False)
        else:
            with path.open("w", encoding="utf-8", newline="") as f:
                pass
        return

    if pd is not None:
        df = pd.DataFrame(rows)
        df.to_csv(path, index=False)
        return

    # Fallback: plain csv
    fieldnames = list(rows[0].keys())
    with path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def write_items_xlsx(path: Path, items: List[Dict[str, Any]]) -> None:
    """
    Write combined items to XLSX if pandas (and a suitable Excel engine)
    is installed. After writing, auto-adjust column widths based on the
    maximum content length per column so headers are readable.
    """
    if pd is None:
        # If pandas isn't available, you can install it or skip XLSX generation.
        return

    rows = [_flatten_item_for_tabular(it) for it in items]
    df = pd.DataFrame(rows)

    path.parent.mkdir(parents=True, exist_ok=True)
    # Let pandas write the basic file first
    df.to_excel(path, index=False)

    # Try to auto-adjust column widths using openpyxl
    try:
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter
    except ImportError:
        # No openpyxl, just keep the default widths
        return

    wb = load_workbook(path)
    ws = wb.active  # pandas writes to the first sheet by default

    from openpyxl.utils import get_column_letter

    # Resize based ONLY on header row (row 1)
    for col_idx, cell in enumerate(ws[1], start=1):
        if cell.value is not None:
            header_len = len(str(cell.value))
        else:
            header_len = 0

        ws.column_dimensions[get_column_letter(col_idx)].width = header_len + 2

    wb.save(path)

